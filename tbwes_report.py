import os
import json
import requests
import numpy as np
import pandas as pd
import timeseries as ts
import app_config as cfg
from datetime import datetime, timedelta, date, time
import time as t
import xlsxwriter
import calendar
import messaging as mg
import openpyxl
from openpyxl import load_workbook,Workbook,drawing,utils
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import datetime as dt
import pytz
from pprint import pprint as pp
import statistics

email = mg.Email()
qr = ts.timeseriesquery()
config = cfg.getconfig()

# unitId = "61caeda654bf6a5bf94bf59a"
tz_IN = pytz.timezone('Asia/Kolkata')   
today = datetime.now(tz_IN)
print(today)
print(today.tzinfo)
endTime = datetime.combine(today, time.min)
endTime = endTime-timedelta(seconds=1)
print(endTime, 'date')
month_start = endTime.replace(day=1, hour = 0, minute=0, second=0)
if today.month < 4:
    year_start_date = month_start.replace(month=4, year = endTime.year-1)
else:
    year_start_date = month_start.replace(month=4)

endTime = 1654387199000 - int(5.5*60*60*1000)
day_start = (endTime - 24*60*60*1000) + 1000
# month_start = 1654041600000 - int(5.5*60*60*1000)
# year_start = 1648771200000 - int(5.5*60*60*1000)
# day_start = calendar.timegm(endTime.timetuple()) * 1000 - 24*60*60*1000 - int(5.5*60*60*1000) + 1000
# month_start = calendar.timegm(month_start.timetuple()) * 1000 - int(5.5*60*60*1000)
# year_start = (calendar.timegm(year_start_date.timetuple()) * 1000) - int(5.5*60*60*1000)
# endTime = ((calendar.timegm(endTime.timetuple()) * 1000) - int(5.5*60*60*1000))
print(endTime, "endtime")
print(day_start, 'daystart')
print(month_start, 'monthstart')
print(year_start, 'yearstart')
startTime = day_start

def get_capacity(unitId):
    url = config["api"]["meta"]+"/units/"+unitId+"/heatrates"
    hr = requests.get(url).json()
    try:
        MAXLOAD = hr[0]["maxLoad"]
    except:
        print("heatrate not found for :"+unitId)
    return MAXLOAD


def get_boilerType(unitId):
    query = {"unitsId":unitId, "equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Flow"}
    urlQuery = config['api']['meta']+'/tagmeta?filter={"where":' + json.dumps(query) + ',"fields":["system"]}'
    response = requests.get(urlQuery)
    if(response.status_code==200):
        system = json.loads(response.content)
        type = system[0]['system']
    else:
        print response.status_code
        print response.content
    return type


def get_incidents(unitId):
    start = (datetime.fromtimestamp((day_start+int(5.5*60*60*1000))/1000).strftime('%Y-%m-%d'))+('T00:00:00')
    end = (datetime.fromtimestamp((day_start+int(5.5*60*60*1000))/1000).strftime('%Y-%m-%d'))+('T23:59:00')
    urlQuery = config['api']['meta']+'/units/' + unitId + '/incidents?filter={"where":{"startTime":{"between":["'+start+'", "'+end+'"]}}, "fields":["priority","startTime","open","id"]}'
    response = requests.get(urlQuery)
    critical = 0
    warning = 0
    deviation = 0
    open = 0
    closed = 0
    if(response.status_code==200):
        incidents = json.loads(response.content)
        for i in incidents:
            if i["priority"] == "danger":
                critical = critical+1
            if i["priority"] == "warning":
                warning = warning+1
            if i["priority"] == "normal":
                deviation = deviation+1
            if i["open"] == True:
                open = open+1
            if i["open"] == False:
                closed =closed+1
    else:
        print response.status_code
        print response.content
    urlQuery = config['api']['meta']+'/units/' + unitId + '/incidents?filter={"where":{"startTime":{"lt":"'+start+'"},"or":[{"endTime":{"gt":"'+start+'"}}, {"open":"true"}]},"fields":["incidentName","startTime","endTime","open"]}'
    response = requests.get(urlQuery)
    if(response.status_code==200):
        prevIncidents = json.loads(response.content)
        prevIncidents = len(prevIncidents)
    return critical+warning+deviation, critical, warning, deviation, open, closed, prevIncidents

def get_tag_uom(unitId, query):
    allTags = get_meta(unitId, query)
    if len(allTags) == 0:
        dataTag = '-'
        UOM = '-'
    else:
        for i in allTags:
            dataTag = [i['dataTagId']]
            UOM = i['measureUnit']
    return dataTag, UOM

def get_meta(unitId, query):
    urlQuery = config['api']['meta']+'/units/' + unitId + '/tagmeta?filter={"where":' + json.dumps(query) + '}'
    response = requests.get(urlQuery)
    if(response.status_code==200):
        allTags = json.loads(response.content)
    else:
        print response.status_code
        print response.content
    return allTags

query1 = [
           {"Steam generated Avg":{"equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Flow"}},
           {"Main steam pressure Avg":{"equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Pressure"}},
           {"Main Steam temperature Avg":{"component":"Steam Circuit", "equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Temperature"}},
           # {"Feed water temperature Avg":{"equipment":"Economizer", "measureProperty":"Feed Water", "measureType":"Temperature", "measureLocation":"Outlet"}},
           {"FD air flow Avg":{"equipment":"Fd Fan", "measureProperty":"Secondary Air", "measureType":"Flow"}},
           {"O2 Avg":{"equipment":"Economizer", "measureProperty":"Flue Gas", "measureType":"O2"}},
           {"ESP inlet temperature Avg":{"equipment":"Air Preheater", "measureProperty":"Flue Gas", "measureType":"Temperature", "measureLocation":"Inlet"}},
           {"Efficiency Avg":{"equipment":"Performance Kpi", "measureProperty":"Boiler", "measureType":"Efficiency"}},
           {"Steam to fuel ratio Avg":{"description":"Steam to Fuel Ratio", "equipment":"Performance Kpi", "measureProperty":"Coal", "measureType":"Ratio"}},
           {"Fuel consumed Avg":{"equipment":"Performance Kpi", "measureProperty":"Coal", "measureType":"Flow"}},
           {"Aux power consumption Avg":{"equipment":"Generator", "measureProperty":"Power", "measureType":"Total Apc"}},
           {"Cost/ton of steam generation":{"equipment":"Performance Kpi", "measureProperty":"Coal", "measureType":"CostPerUnit"}}]
query2 = [{"Total steam generated/Day":{"equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Flow"}},
           {"Total fuel consumed/Day":{"equipment":"Performance Kpi", "measureProperty":"Coal", "measureType":"Flow"}},
           {"Total Aux power consumption day avg":{"equipment":"Generator", "measureProperty":"Power", "measureType":"Total Apc"}},
           {"Availability":{"equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Flow"}}
           ]
queries = query1+query2

steam_flow = {"equipment":"Superheater", "measureProperty":"Main Steam", "measureType":"Flow"}

def get_data(tag, startTime, endTime):
    qr.addMetrics(tag)
    qr.chooseTimeType("absolute",{"start_absolute":startTime, "end_absolute":endTime})
    qr.addAggregators([{"name":"avg", "sampling_value":1,"sampling_unit":'minutes'}])
    qr.submitQuery()
    qr.formatResultAsDF()
    try:
        values = qr.resultset["results"][0]['data']
    except:
        values = pd.DataFrame()
    return(values)



def getMinMaxAvg(tag, startTime, endTime):
    try:
        day_data = get_data(tag, startTime, endTime)
        day_data = day_data.set_index('time')
        
        minimum = ((day_data[tag].min()).values[0]).round(2)
        maximum = ((day_data[tag].max()).values[0]).round(2)
        average = ((day_data[tag].mean()).values[0]).round(2)
    except:
        minimum = "-"
        maximum = "-"
        average = "-"
    return minimum, maximum, average


def get_run_hours(steam_tag, startTime, endTime, threshold):
    df = get_data(steam_tag, startTime, endTime)
    try:
        df = df[df[steam_tag[0]]>threshold]
        df_array = np.array(df['time'])
        diff_ary = (df_array[1:] - df_array[:-1])/60000
        diff, = np.where(diff_ary>15)
        diff = diff+1
        hours = 0
        if len(diff) == 0:
            hrs = round((df['time'].iloc[-1] - df['time'].iloc[0])/(3600000*1.0))
        else:
            for ii in range(len(diff)):
                if ii == 0:
                    start = df['time'][0]
                else:
                    start = df['time'][diff[ii-1]]
                end = df['time'][diff[ii]-1]
                hours=hours+((end - start)/(3600000*1.0))
            start = df['time'][diff[-1]]
            
            end = df['time'].iloc[-1]
            hrs = round(hours+((end - start)/(3600000*1.0)), 2)
        if np.isnan(hrs) == True:
            hrs = "-"
    except:
        hrs = "-"
    return hrs


def get_total(tag_dataTag, startTime, endTime):
    try:
        data = get_data(tag_dataTag, startTime, endTime)
        data = data.set_index('time')
        total = (data.mean()).values[0]
    except:
        total = "-"
    return total
dict2 = []
def getReport():
    unit_ids = []
    query = {"bu":"tbwes"}
    urlQuery = config['api']['meta']+'/customers/?filter={"where":' + json.dumps(query) + '}'
    response = requests.get(urlQuery)
    if(response.status_code==200):
        allCustomers = json.loads(response.content)
        # print(allCustomers)
        for customer in allCustomers:
            print(customer['name'])
            query = {"customerId":customer["id"]}
            urlQuery = config['api']['meta']+'/units/?filter={"where":' + json.dumps(query) + '}'
            response = requests.get(urlQuery)
            if(response.status_code==200):
                allUnits = json.loads(response.content)
                print(allUnits)
                
                jj = 0
                finaldf = pd.DataFrame()
                finaldf2 = pd.DataFrame()
                
                index2 = []
                index2.append(("Description", "Boiler No/Name", "Capacity", "Boiler Type", "KEY OPERATION PARAMETERS", ""))
                index2.append(("Unit", "No/Name", "TPH", "Type", "", ""))
                for unit in allUnits:
                    if "status" in unit:
                        blrdetail = []
                        # incdntdetail = []
                        unitId = unit['id']
                        print(unitId, unit['name'])
                        boilerName = unit['name']
                        capacity = get_capacity(unitId)
                        boilerType = get_boilerType(unitId)
                        index = []
                        
                        index.append(("Boiler Details", boilerName, capacity, boilerType, "", "MIN"))
                        index.append(("Boiler Details", boilerName, capacity, boilerType, "", "MAX"))
                        index.append(("Boiler Details", boilerName, capacity, boilerType, "", "AVG"))
                        
                        steam_tags, u = get_tag_uom(unitId, steam_flow)
                        threshold = 5
                        runTime = get_run_hours(steam_tags, startTime, endTime, threshold)
                        print(runTime, "RUNTIME")
                        print("*"*20)
                        
                        mini = []
                        maxi = []
                        avgi = []
                        desc = []
                        uom = []
                        for dict1 in queries:
                            query = dict1.values()[0]
                            tag, UOM = get_tag_uom(unitId, query)
                            
                            total_hours = ((endTime+1000) - startTime)/(60*60*1000*1.0)
                            
                            
                            if dict1.keys()[0] in ["Total steam generated/Day", "Total Aux power consumption day avg", "Total fuel consumed/Day"]:
                                print(tag)
                                total = get_total(tag, startTime, endTime)
                                print(total, dict1.keys()[0])
                                try:
                                    total = (total*runTime).round(2)
                                except:
                                    total = "-"
                                mini.append(total)
                                maxi.append(total)
                                avgi.append(total)
                                desc.append(dict1.keys()[0])
                                uom.append(UOM)
                            elif dict1.keys()[0] in ['Availability']:
                                print(total_hours, "total hours")
                                print(runTime, "runtime")
                                try:
                                    total = ((runTime/(total_hours*1.0))*100).round(2)
                                except:
                                    total = "-"
                                mini.append(total)
                                maxi.append(total)
                                avgi.append(total)
                                desc.append(dict1.keys()[0])
                                uom.append(UOM)
                            else:
                                minimum, maximum, average = getMinMaxAvg(tag, startTime, endTime)
                                mini.append(minimum)
                                maxi.append(maximum)
                                avgi.append(average)
                                desc.append(dict1.keys()[0])
                                uom.append(UOM)
                        
                        totalIncidents, critical, warning, deviation, open, closed, prevIncidents = get_incidents(unitId)
                        dict2 = [{"No of incidents generated":totalIncidents}, {"Critical incidents":critical}, {"Warning incidents":warning}, {"Deviation incidents":deviation}, {"No of open incidents":open}, {"No of closed incidents":closed}, {"Previous day open incidents":prevIncidents}]
                        for j in dict2:
                            incdntValues = j.values()[0]
                            incdntDesc = j.keys()[0]
                            mini.append(incdntValues)
                            maxi.append(incdntValues)
                            avgi.append(incdntValues)
                            desc.append(incdntDesc)
                            uom.append("")
                        
                        index = pd.MultiIndex.from_tuples(index)
                        list_tuples = list(zip(mini, maxi, avgi))
                        df = pd.DataFrame(data=list_tuples, columns=index)
                        finaldf = pd.concat([finaldf, df], axis = 1)
                        finaldf = finaldf.sort_index(axis=1)
                        
                index2 = pd.MultiIndex.from_tuples(index2)
                list_tuples2 = list(zip(desc, uom))
                df2 = pd.DataFrame(data=list_tuples2, columns=index2)
                finaldf2 = pd.concat([df2, finaldf], axis = 1)
                print(finaldf2)
                if finaldf.empty:
                    continue
                writer = pd.ExcelWriter("/space/es-master/src/RaviRanjan/TBWESReports/"+customer['name']+" Daily Customer Report Report.xlsx", engine='openpyxl')
                finaldf2.to_excel(writer, index=True, sheet_name='Report', startrow = 1, startcol = 0)
                worksheet = writer.sheets['Report']
                dflen = len(query1)+len(query2)+len(dict2)+8
                for row in worksheet['A1:A30']:
                  for cell in row:
                    cell.value = None
                
                incidentLineNo = len(query1)+len(query2)+9
                worksheet.insert_rows(incidentLineNo)
                worksheet.merge_cells(start_row=incidentLineNo, start_column=2, end_row=incidentLineNo, end_column=finaldf2.shape[1]+1)
                incidentTitle = worksheet.cell(row=incidentLineNo, column=2)
                incidentTitle.value = 'INCIDENTS'
                incidentTitle.alignment = Alignment(horizontal='left', vertical='center')
                incidentTitle.font = Font(name='Calibri',size=15,bold=True,color='000000')
                incidentTitle.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                start = len(query1)+9
                end = len(query1)+len(query2)+len(dict2)+10
                for x in range(start,end):
                    for y in range(4,finaldf2.shape[1],3):
                        worksheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y+2)
                
                dims = {}
                for row in worksheet.rows:
                    for cell in row:
                        if cell.value:
                            try:
                                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                            except:
                                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(cell.value.encode('utf-8'))))
                for col, value in dims.items():
                    worksheet.column_dimensions[col].width = value + 5
                
                bd = Side(style='thin', color="000000")
                for cell in worksheet._cells.values():
                    cell.font = Font(name='Calibri',size=12, bold = True)
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=finaldf2.shape[1]-3)
                title = worksheet.cell(row=1, column=2)
                title.value = 'DAILY OPERATION REPORT'
                title.alignment = Alignment(horizontal='left', vertical='center')
                title.font = Font(name='Calibri',size=15,bold=True,color='000000')
                title.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                yesterday = (datetime.fromtimestamp((endTime+int(5.5*60*60*1000))/1000)).strftime('%d%b%Y')
                print(yesterday)
                worksheet.merge_cells(start_row=1, start_column=finaldf2.shape[1]-2, end_row=1, end_column=finaldf2.shape[1]+1)
                date = worksheet.cell(row=1, column=finaldf2.shape[1]-2)
                date.value = "Date - "+yesterday
                date.alignment = Alignment(horizontal='center', vertical='center')
                date.font = Font(name='Calibri',size=15,bold=True,color='000000')
                date.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                for row in worksheet.iter_rows(min_row=2, max_row=2, min_col=2,max_col=finaldf2.shape[1]):
                    for cell in row:
                        cell.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                # worksheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=finaldf2.shape[1])
                for row in worksheet.iter_rows(min_row=6, max_row=6, min_col=2,max_col=finaldf2.shape[1]+1):
                    for cell in row:
                        cell.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                for row in worksheet.iter_rows(min_row=3, max_row=finaldf2.shape[0]+9, min_col=2,max_col=3):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                
                worksheet.merge_cells(start_row=1, start_column=finaldf2.shape[1]+2, end_row=5, end_column=finaldf2.shape[1]+2)

                for row in worksheet.iter_rows(min_row=1, max_row=finaldf2.shape[0]+9, min_col=finaldf2.shape[1]+2,max_col=finaldf2.shape[1]+2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                
                row_number = str(1)
                col_idx = chr(64+finaldf2.shape[1]+2)
                cellNo = col_idx+row_number
                worksheet.column_dimensions[col_idx].width = value + 40
                
                remarks = worksheet.cell(row=6, column=finaldf2.shape[1]+2)
                remarks.value = "Remarks"
                remarks.alignment = Alignment(horizontal='center', vertical='center')
                remarks.font = Font(name='Calibri',size=12,bold=True,color='000000')
                remarks.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                img = openpyxl.drawing.image.Image("/space/es-master/src/RaviRanjan/TBWESReports/edgelivelogo.png")
                worksheet.add_image(img, cellNo)
                
                for row in worksheet.iter_rows(min_row=end, max_row=end+8, min_col=2,max_col=finaldf2.shape[1]+2):
                    for cell in row:
                        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                worksheet.merge_cells(start_row=end, start_column=2, end_row=end, end_column=finaldf2.shape[1]+2)
                observation = worksheet.cell(row=end, column=2)
                observation.value = 'Observations & Recommendations'
                observation.alignment = Alignment(horizontal='center', vertical='center')
                observation.font = Font(name='Calibri',size=12,bold=True,color='000000')
                observation.fill = PatternFill(fill_type='solid',start_color="FFFF00",end_color="FFFF00")
                
                worksheet.merge_cells(start_row=end+1, start_column=2, end_row=end+8, end_column=finaldf2.shape[1]+2)
                writer.save()
                
                msg_body = """<h3>Daily Report Attached</h3>"""
                body = {
                    "to": ["ravi.r@exactspace.co"],
                    "subject": "Daily Customer Report",
                    "html": msg_body,
                    "f1": "/space/es-master/src/RaviRanjanTBWESReports/"+customer['name']+" Daily Customer Report Report.xlsx",
                    "cc":[], # Not mandatory #list of email ID strings, when needed
                    "bcc":[] # Not mandatory #list of email ID strings, when needed
                }
                print("mailing")
                email.sendSESMailWithAttach(body)

                # os.remove("/space/es-master/src/RaviRanjan/TBWESReports/"+customer['name']+" Daily Customer Report Report.xlsx")
    else:
        print response.status_code
        print response.content
    return finaldf
report = getReport()