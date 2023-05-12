import requests, json
import pandas as pd
import sys
import datetime
import app_config as cfg
config = cfg.getconfig()
import timeseries as ts
qr = ts.timeseriesquery()
import openpyxl
from openpyxl import load_workbook,Workbook,drawing,utils
import warnings
URL = config["api"]["meta"]
print URL
import xlsxwriter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import time
from datetime import timedelta
from datetime import date


import messaging as mg

email = mg.Email()

# UNIT_ID = sys.argv[1]
UNIT_ID='61dd822329c9e07656414708'
dataConfigFileName = str(UNIT_ID) + "_2.json"
print dataConfigFileName

f = open(dataConfigFileName)
dataConfig = json.load(f)
f.close()


def getAverage(tagList, startTime, endTime):
    qr = ts.timeseriesquery()
    qr.addMetrics(tagList)
    qr.chooseTimeType("date",{"start_absolute":str(startTime), "end_absolute":str(endTime)})
#     qr.chooseTimeType("date",{"start_absolute":"01-01-2020 00:00", "end_absolute":"31-05-2020 23:59"})
    qr.addAggregators([{"name":"avg", "sampling_value":2,"sampling_unit":"days"}])
    qr.submitQuery()
    qr.formatResultAsDF()
    if (len(qr.resultset["results"]) > 0):
        values = qr.resultset["results"][0]["data"]
        values = values.round(3)
        return values
    else:
        #print "no datapoints found in production kairos"
        return pd.DataFrame()
        
def getMax  (tagList, startTime, endTime):
    qr = ts.timeseriesquery()
    qr.addMetrics(tagList)
    qr.chooseTimeType("date",{"start_absolute":str(startTime), "end_absolute":str(endTime)})
#     qr.chooseTimeType("date",{"start_absolute":"01-01-2020 00:00", "end_absolute":"31-05-2020 23:59"})
    qr.addAggregators([{"name":"max", "sampling_value":2,"sampling_unit":"days"}])
    qr.submitQuery()
    qr.formatResultAsDF()
    if (len(qr.resultset["results"]) > 0):
        values = qr.resultset["results"][0]["data"]
        values = values.round(3)
        return values
    else:
        #print "no datapoints found in production kairos"
        return pd.DataFrame()
        
        
def getMin(tagList, startTime, endTime):
    qr = ts.timeseriesquery()
    qr.addMetrics(tagList)
    qr.chooseTimeType("date",{"start_absolute":str(startTime), "end_absolute":str(endTime)})
#     qr.chooseTimeType("date",{"start_absolute":"01-01-2020 00:00", "end_absolute":"31-05-2020 23:59"})
    qr.addAggregators([{"name":"min", "sampling_value":2,"sampling_unit":"days"}])
    qr.submitQuery()
    qr.formatResultAsDF()
    if (len(qr.resultset["results"]) > 0):
        values = qr.resultset["results"][0]["data"]
        values = values.round(3)
        return values
    else:
        #print "no datapoints found in production kairos"
        return pd.DataFrame()

# def getMinMaxAvg(tagsList):
    # print tagsList
    # return pd.DataFrame
  
def getdifference(tagList, startTime, endTime):
    qr = ts.timeseriesquery()
    qr.addMetrics(tagList)
    qr.chooseTimeType("date",{"start_absolute":str(startTime), "end_absolute":str(endTime)})
#     qr.chooseTimeType("date",{"start_absolute":"01-01-2020 00:00", "end_absolute":"31-05-2020 23:59"})
    # qr.addAggregators([{"name":"min", "sampling_value":2,"sampling_unit":"days"}])
    qr.submitQuery()
    qr.formatResultAsDF()
    if (len(qr.resultset["results"]) > 0):
        values = qr.resultset["results"][0]["data"]
        values = values[tagList[0]][len(values)-1]-values[tagList[0]][0]
        return values
    else:
        #print "no datapoints found in production kairos"
        return pd.DataFrame()
    
    
    
    
    
def getTagFromQuery(query):
    if "dataTagId" in query:
        return query["dataTagId"]
    url = URL + '/units/' +str(UNIT_ID) +'/tagmeta?filter={"where":'+json.dumps(query)+', "fields":["dataTagId"]}'
    res = requests.get(url)
    if res.status_code == 200:
        datatag = json.loads(res.content)
        if datatag:
            return datatag[0]["dataTagId"]
        else:
            print "\t\tMeta Query returning empty tags.Kindly check list of queries once and this specific query once.", str(query)
    else:
        print "Requests Error in function getTagFromQuery"
        return ""


def getTagsFromQueries(queriesList):
    tags = []
    for query in queriesList:
        tag = getTagFromQuery(query)
        if tag:
            tags.append(tag)
        else:
            continue
            
    return tags
    
def getDatePriorFromToday(NUMBER_OF_DAYS):
    return datetime.datetime.strftime((datetime.date.today() - datetime.timedelta(days=NUMBER_OF_DAYS)), "%d-%m-%Y") + " 07:00"


def getDurationList(startDate):
    durationList = []
    totalSteps = NUMBER_OF_DAYS * 3
    for i in range(totalSteps):
        firstStep = datetime.datetime.strftime((datetime.datetime.strptime(startDate, "%d-%m-%Y %H:%M") + datetime.timedelta(hours=(i*8))), "%d-%m-%Y %H:%M")
        secondStep = datetime.datetime.strftime((datetime.datetime.strptime(startDate, "%d-%m-%Y %H:%M") + datetime.timedelta(hours=((i+1)*8))), "%d-%m-%Y %H:%M")
        addToList = [firstStep, secondStep]
        durationList.append(addToList)
    return durationList


def Average(lst):
    return sum(lst) / len(lst)
    
    
    
NUMBER_OF_DAYS = 7

startDate = getDatePriorFromToday(NUMBER_OF_DAYS)
durationList = getDurationList(startDate)
print len(durationList)

def getDailySet(tag, dur1, dur2):
    df = getAverage([tag], dur1, dur2)
    df.round(3)
    if df.empty:
        dd[tag + "__avg"].append("No Data")
    else:
        dd[tag + "__avg"].append(df[tag].values[0])
    
results = []
lavgtags = []
lmmatags = []
stendtags=[]
for k, v in dataConfig.items():
    print k
    for k2,v2 in dataConfig[k].items():
            for k3, v3 in dataConfig[k][k2].items():
                if k2 == "avg":
                    tags = getTagsFromQueries(v3)
                    for tag in tags:
                        dd = {}
                        dd[tag + "__avg"] = []
                        salCounter = 0
                        for duration in durationList:

                            #vals = getDailySet(tag, duration[0], duration[1])
                            df = getAverage([tag], duration[0], duration[1])
                            df.round(3)
                            if df.empty:
                                dd[tag + "__avg"].append("No Data")
                            else:
                                dd[tag + "__avg"].append(float(df[tag].values[0]))
                            #print dd[tag + "__avg"], len(dd[tag + "__avg"]), salCounter
                            salCounter = salCounter + 1
                            if salCounter % 3 == 0:
                                #print dd[tag + "__avg"][-3:], "%%%%%%%", salCounter
                                if all(isinstance(ele, float)  for ele in dd[tag + "__avg"][-3:]):
                                    dd[tag + "__avg"].append(sum(dd[tag + "__avg"][-3:]))
                                else:
                                    dd[tag + "__avg"].append("No Data")
                                #print dd[tag + "__avg"], "%%%%%%%%%%%%%%%%%%%%", len(dd[tag + "__avg"])
                        # print "\n\n"
                        print dd
                        lavgtags.append(dd)
                        print "\n\n"
                    #results.append(df.to_dict(orient='records'))
                elif k2 == "minMaxAvg":
                    tags = getTagsFromQueries(v3)
                    for tag in tags:
                        salCounter01 = 0
                        salCounter02 = 0
                        salCounter03 = 0
                        dd = {}
                        dd[tag + "__avg"] = []
                        dd[tag + "__min"] = []
                        dd[tag + "__max"] = []
                        for duration in durationList:
                            dfmin = getMin([tag], duration[0], duration[1])
                            if dfmin.empty:
                                dd[tag + "__min"].append("No Data")
                            else:
                                dd[tag + "__min"].append(dfmin[tag].values[0])
                            salCounter01 = salCounter01 + 1
                            if salCounter01 % 3 == 0:
                                if all(isinstance(ele, float)  for ele in dd[tag + "__min"][-3:]):
                                    dd[tag + "__min"].append(Average(dd[tag + "__min"][-3:]))
                                else:
                                    dd[tag + "__min"].append("No Data")
                            dfmax = getMax([tag], duration[0], duration[1])
                            if dfmax.empty:
                                dd[tag + "__max"].append("No Data")
                            else:
                                dd[tag + "__max"].append(dfmax[tag].values[0]) 
                            salCounter02 = salCounter02 + 1
                            if salCounter02 % 3 == 0:
                                if all(isinstance(ele, float)  for ele in dd[tag + "__max"][-3:]):
                                    dd[tag + "__max"].append(Average(dd[tag + "__max"][-3:]))
                                else:
                                    dd[tag + "__max"].append("No Data")                 
                            df = getAverage([tag], duration[0], duration[1])
                            if df.empty:
                                dd[tag + "__avg"].append("No Data")
                            else:
                                dd[tag + "__avg"].append(df[tag].values[0])
                            salCounter03 = salCounter03 + 1
                            if salCounter03 % 3 == 0:
                                if all(isinstance(ele, float)  for ele in dd[tag + "__avg"][-3:]):
                                    dd[tag + "__avg"].append(Average(dd[tag + "__avg"][-3:]))
                                else:
                                    dd[tag + "__avg"].append("No Data")

                            #dfmin["date"] = pd.to_datetime(dfmin["time"], unit='ms')
                            #dfmax["date"] = pd.to_datetime(dfmax["time"], unit='ms')
                            #df["date"] = pd.to_datetime(df["time"], unit='ms')
                            #results.append(dfmin.to_dict(orient='records'))
                            #results.append(dfmax.to_dict(orient='records'))
                            #results.append(df.to_dict(orient='records'))
                        print "\n\n"
                        print dd, len(dd[tag + "__avg"]),len(dd[tag + "__min"]),len(dd[tag + "__max"])
                        lmmatags.append(dd)
                                # print "\n\n"
                elif k2 == "startend":
                    tags = getTagsFromQueries(v3)
                    for tag in tags:
                        dd = {}
                        dd[tag + "__startend"] = []
                        salCounter = 0
                        for duration in durationList:

                            #vals = getDailySet(tag, duration[0], duration[1])
                            df_t = getdifference([tag], duration[0], duration[1])
                            df_t.round(3)
                            try:
                                if df_t.empty:
                                    dd[tag + "__startend"].append("No Data")
                            except:
                                dd[tag + "__startend"].append(float(df_t))
                            #print dd[tag + "__startend"], len(dd[tag + "__startend"]), salCounter
                            salCounter = salCounter + 1
                            if salCounter % 3 == 0:
                                #print dd[tag + "__startend"][-3:], "%%%%%%%", salCounter
                                if all(isinstance(ele, float)  for ele in dd[tag + "__startend"][-3:]):
                                    dd[tag + "__startend"].append(sum(dd[tag + "__startend"][-3:]))
                                else:
                                    dd[tag + "__startend"].append("No Data")
                                #print dd[tag + "__startend"], "%%%%%%%%%%%%%%%%%%%%", len(dd[tag + "__startend"])
                        # print "\n\n"
                        # print dd
                        stendtags.append(dd)
                
                
                else:
                    print "\t\tConfig Inaccurate. contact Developer."
                    
print durationList

# 4,6,19,81
# 5,7,25,72
def getTagsList(sheet,i,a1,a2,s1,s2,m1,m2):
    print s1, s2
    pfix = 'HRD_'
    sfix = '__avg'
    
    savg='__avg'
    smin='__min'
    smax='__max'
    stend='__startend'
    mmatags={}
    avgtags={}
    for k in range(int(a1),int(a2)):
        x=sheet["D"+str(k)].value
        if (x != None and x!='na.'):
            etag=pfix + x.replace("-","_")+sfix
            for val in lavgtags:
                if etag in val:
                    for cel in range(len(charList)):
                        sheet[charList[cel]+str(k)].value=val[etag][cel]
                        # sheet["G"+str(k)].value=val[etag][1]
                        # sheet["H"+str(k)].value=val[etag][2]
                        # sheet["J"+str(k)].value=val[etag][3]
                        # sheet["K"+str(k)].value=val[etag][4]
                        # sheet["L"+str(k)].value=val[etag][5]
                        # sheet["N"+str(k)].value=val[etag][6]
                        # sheet["O"+str(k)].value=val[etag][7]
                        # sheet["P"+str(k)].value=val[etag][8]
    for z in range(int(s1),int(s2)):
        xt=sheet["D"+str(z)].value
        print xt, "#####33333"
        if (xt != None and xt!='na.'):
            etag=pfix + xt.replace("-","_")+stend
            print etag, "$$$$$$44"
            for val in stendtags:
                # print val, "############"
                if etag in val:
                    print etag, val, "$$$$$$$$$$$$$$$$"
                    for cel in range(len(charList)):
                        sheet[charList[cel]+str(z)].value=val[etag][cel]
                    
    for j in range(int(m1),int(m2),3):
        print j 
        y=sheet["D"+str(j)].value
        print y 
        if (y != None and y!='na.'):
            mintag=pfix + y.replace("-","_")+smin
            maxtag=pfix + y.replace("-","_")+smax
            avgtag=pfix + y.replace("-","_")+savg
            
            print mintag, maxtag, avgtag
            for val in lmmatags:
               
                if mintag in val:
                    for cel in range(len(charList)):
                        sheet[charList[cel]+str(j)].value=val[mintag][cel]
                    # sheet["G"+str(j)].value=val[mintag][1]
                    # sheet["H"+str(j)].value=val[mintag][2]
                    # sheet["J"+str(j)].value=val[mintag][3]
                    # sheet["K"+str(j)].value=val[mintag][4]
                    # sheet["L"+str(j)].value=val[mintag][5]
                    # sheet["N"+str(j)].value=val[mintag][6]
                    # sheet["O"+str(j)].value=val[mintag][7]
                    # sheet["P"+str(j)].value=val[mintag][8]

                if maxtag in val:
                    for cel in range(len(charList)):
                        sheet[charList[cel]+str(j+1)].value=val[maxtag][cel]
                    # sheet["F"+str(j+1)].value=val[maxtag][0]
                    # sheet["G"+str(j+1)].value=val[maxtag][1]
                    # sheet["H"+str(j+1)].value=val[maxtag][2]
                    # sheet["J"+str(j+1)].value=val[maxtag][3]
                    # sheet["K"+str(j+1)].value=val[maxtag][4]
                    # sheet["L"+str(j+1)].value=val[maxtag][5]
                    # sheet["N"+str(j+1)].value=val[maxtag][6]
                    # sheet["O"+str(j+1)].value=val[maxtag][7]
                    # sheet["P"+str(j+1)].value=val[maxtag][8]

                if avgtag in val:
                    for cel in range(len(charList)):
                        sheet[charList[cel]+str(j+2)].value=val[avgtag][cel]
                    # sheet["F"+str(j+2)].value=val[avgtag][0]
                    # sheet["G"+str(j+2)].value=val[avgtag][1]
                    # sheet["H"+str(j+2)].value=val[avgtag][2]
                    # sheet["J"+str(j+2)].value=val[avgtag][3]
                    # sheet["K"+str(j+2)].value=val[avgtag][4]
                    # sheet["L"+str(j+2)].value=val[avgtag][5]
                    # sheet["N"+str(j+2)].value=val[avgtag][6]
                    # sheet["O"+str(j+2)].value=val[avgtag][7]
                    # sheet["P"+str(j+2)].value=val[avgtag][8]
    dlist=[]
    def datesofpastsevendays():

        today=date.today()
        yes = date.today()-timedelta(days = 1)
        for x in range(7):
            dlist.append((yes - timedelta(days=x)).strftime("%d-%m-%Y"))
        
        
        return dlist
    

    dateslist=datesofpastsevendays()
           
    clist=["E","I","M","Q","U","Y","AC"]
    for cel in range(len(clist)):
        if "TS 1" in i:
            print(cel)
          
            print(clist[cel],sheet[clist[cel]+str(2)].value)
            print(sheet[clist[cel]+str(2)].value)
            sheet[clist[cel]+str(2)].value= (dateslist)[6-cel]
            sheet[clist[cel]+str(20)].value=(dateslist)[6-cel]
            
        elif "TS 3" in i:
            print(cel)
          
            print(clist[cel],sheet[clist[cel]+str(2)].value)
            print(sheet[clist[cel]+str(2)].value)
            sheet[clist[cel]+str(2)].value= (dateslist)[6-cel]
            sheet[clist[cel]+str(23)].value=(dateslist)[6-cel]

            
       
        
    return 
    

charList = ["E","F",	"G",	"H",	"I",	"J",	"K",	"L",	"M",	"N",	"O",	"P",	"Q",	"R",	"S",	"T",	"U",	"V",	"W",	"X",	"Y",	"Z",	"AA",	"AB",	"AC",	"AD",	"AE",	"AF"]
FinalRep="/space/es-master/src/wws_reports/wws_reports_3.xlsx"
wb = load_workbook(FinalRep)
g=wb.sheetnames
for i,j in zip(g,range(len(g))):
    print i, "%%%%%%%%%%%%%%%%%%%%%%%%%"
    sheet=wb.worksheets[j]
    if "TS 1" in i:
        getTagsList(sheet,i,4,6,4,15,19,81)
    elif "TS 3" in i:
        getTagsList(sheet,i,5,7,8,21,25,72)
    else:
        print ("please check the source file")
# print("Updating Excel")
wb.save(FinalRep)

wb = load_workbook(FinalRep)

emails=["shravya.k@exactspace.co","jason.d@exactspace.co","rishi.hiran@thermaxglobal.com","amit.kulkarni@thermaxglobal.com","atish.kulkarni@thermaxglobal.com","Dattatraya.Bhoskar@thermaxglobal.com","Pravin.Rangole@thermaxglobal.com","Sauraj.Gill@thermaxglobal.com","K.Madhavan@thermaxglobal.com","Mandar.Erande@thermaxglobal.com","abhijeet.herlekar@Thermaxglobal.com","Ujwal.Dambe@thermaxglobal.com","vishal.mehra@Thermaxglobal.com","Ravikiran.Lad@thermaxglobal.com","Shankarrao.Borate@thermaxglobal.com","Rathin.Chowdhury@thermaxglobal.com","Sambhaji.Thorat@Thermaxglobal.com","nandan.prabhune@thermaxglobal.com","Sayali.Jadhav@thermaxglobal.com", "Rravindar.Verma@pepsico.com","Sagar.Tingare@thermaxglobal.com"]
#emails=["shravya.k@exactspace.co","bibhudatta.s@exactspace.co"]
body = {
        "to": emails,
        "subject": "Pepsi Co Scada Report " + date.today().strftime("%d-%m-%Y"),
        "html":"<h3> </h3>",
        "f1": FinalRep,
        "f2":"", # Not mandatory
        "f3":"" # Not mandatory
        }
print("mailing")
email.sendSESMailWithAttach(body)