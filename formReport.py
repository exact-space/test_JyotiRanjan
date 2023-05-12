import requests
import json
import urllib
from collections import defaultdict
import pandas as pd
from datetime import datetime
import numpy as np
import time
import xlsxwriter
import timeseries as ts
import os
import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
import time as tm
from openpyxl import load_workbook,Workbook,drawing,utils
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from datetime import datetime, time,timedelta
import calendar
from matplotlib.pyplot import figure
import warnings
warnings.filterwarnings("ignore")
from datetime import timedelta
from datetime import date

figure(figsize=(18, 5), dpi=80)
import messaging as mg

email = mg.Email()

#unitId="63be6ce733b4cc000725c296" #meghamani
#unitId="611b7fd57afa992d36e308b0" #JKLC

def ExcelReader():

    url_loopback='https://pulse.thermaxglobal.com/exactapi'
    def getToken(url):
        query= url+ '/Users/login'
        body={"email":"arun@exactspace.co","password":"Thermax@123","ttl":0}
        res=requests.post(url=query, json=body)
        token=json.loads(res.content)
        token=token['id']
    #     print token
        return token

    token=getToken(url_loopback)

    def getAllUnits():
        url = url_loopback + "/units?&access_token=" + token
        response = requests.get(url)
        if response.status_code == 200:
            return json.loads(response.content)

        else:
            print response.status_code
            print "Some error in fetching units from DB"

    def filterUnits(allUnitIds, businessUnit):
        buid=[]
        for i in allUnitIds:
            if "bu" in i:
                if businessUnit in i["bu"]:
                    buid.append(i["id"])
        return buid 

    def getTagsForEachUnit(unitId):
        url = url_loopback + '/tagmeta?filter={"where":{"unitsId":"'+ unitId +'","equipmentName":"Thermread Plus 1"},"fields":["dataTagId","measureProperty","measureType","measureUnit"]}'
        response = requests.get(url,headers={'Authorization': token})
        if response.status_code == 200:
            tags=json.loads(response.content)
            return tags
        else:
            print response.status_code
            print "Some error in fetching tags from DB"
        return

    allUnits = getAllUnits()
    #print("allUnits",allUnits)
    businessUnit = "chemicals"
    
    def todayyesterday():
        yesterday = datetime.strftime(date.today() -timedelta(days = 1), "%d-%m-%Y %H:%M" )
        today =datetime.strftime(date.today(), "%d-%m-%Y %H:%M")
        today = datetime.strftime(datetime.strptime(today, "%d-%m-%Y %H:%M") - timedelta(minutes = 1), "%d-%m-%Y %H:%M")
        
        #convert time into epoch
        def convert(time):
            val=""
            year=int(time[6:10])
            month=int(time[3:5])
            day=int(time[0:2])
            hour=int(time[11:13])
            mins=int(time[14:])
            val=int(datetime(year, month, day, hour, mins).strftime('%s'))
            return int(val*1000)
        
        #calling convert fxn to change the date format to epoc
        yesterday=convert(yesterday)
        today=convert(today)
        
        #returning values
        return yesterday,today

    #calling the fxn
    yesterday,today= todayyesterday()

    UnitsForSpecificBusinessUnits = filterUnits(allUnits, businessUnit)
    # TagsForQueryAndUnit=defaultdict(dict)
    
    for unitid in ["63be6ce733b4cc000725c296"]:#UnitsForSpecificBusinessUnits:
        TagsForQueryAndUnit=[]
        var=getTagsForEachUnit(unitid)
        print("varrrrr",var)
        tags=[]
        if var:
            for i in var:
                tags.append(i['dataTagId'])
            print ("tags",tags)
        else:
            tags = []
        
        unitName=tags[0].split("_")[0]           
        print ("unnnnit",unitName)
        # print tags
        for i in tags:
            d= {
            "metrics": [
            {
            "tags": {},
            "name": i,
            "aggregators": [
            {
            "name": "first",
            "sampling": {
            "value": "30",
            "unit": "minutes"
            }
            }
            ]
            }
            ],
            "plugins": [],
            "cache_time": 0,
            "start_absolute": yesterday,
            "end_absolute": today
            }
    #         print d-
            TagsForQueryAndUnit.append(d)
        
        # print(TagsForQueryAndUnit)
        dfo=pd.DataFrame()
        columns=["pH","ORP (mV)","CONDUCTIVITY "+ u'\u00B5' + "s/cm","Temp " + u'\u00B0'+"C","FLUOROMETER (PPM)","CORROSION (MPY)","RELAY1","RELAY2","RELAY3","RELAY4","RELAY5","RELAY6","Flowswitch"]
        for d,i in zip(TagsForQueryAndUnit,columns):
        #     print (d['metrics'][0]['name'])
            url_kairos="https://pulse.thermaxglobal.com/kairosapi/api/v1/datapoints/query"
            headers={'Authorization': token}
            res=requests.post(json=d,headers=headers,url=url_kairos)
            value=json.loads(res.content)
            print ("chlo chle",value)
            df=pd.DataFrame(value["queries"][0]["results"][0]["values"],columns=["Time",i])
            # print df.head()
            # print df.columns
            dfo=pd.concat([dfo,df],axis=1,sort=False)
            print("dfo",dfo)
        df_final = dfo.T.groupby(level=0).first().T
        
        first_column = df_final.pop('pH')
        df_final.insert(0, 'pH', first_column)
    #     df_final.drop(['Temp'],axis=1,inplace=True)
        first_column = df_final.pop('Time')
        df_final.columns
        df_final.insert(0, 'Time', first_column)
        df_final["pH within specs"] = pd.cut(df_final["pH"],[7.5,8.5], precision=3, labels=["Y"])
        df_final["ORP within specs"] = pd.cut(df_final["ORP (mV)"],[-1500,1500], precision=3, labels=["Y"])
        df_final["CONDUCTIVITY within specs"] = pd.cut(df_final["CONDUCTIVITY "+ u'\u00B5' + "s/cm"],[300,900], precision=3, labels=["Y"])
        df_final["CORROSION within specs"] = pd.cut(df_final["CORROSION (MPY)"],[0,3], precision=3, labels=["Y"])
        df_final["FLUOROMETER within specs"] = pd.cut(df_final["FLUOROMETER (PPM)"],[20,20], precision=3, labels=["Y"])
        df_final['Flowswitch']=df_final['Flowswitch'].apply(lambda x: "Flow" if x==1 else "No Flow")
        df_final['RELAY1']=df_final['RELAY1'].apply(lambda y:  "On" if y==1  else "Off")
        df_final['RELAY2']=df_final['RELAY2'].apply(lambda x2: "On" if x2==1 else "Off")
        df_final['RELAY3']=df_final['RELAY3'].apply(lambda x3: "On" if x3==1 else "Off")
        df_final['RELAY4']=df_final['RELAY4'].apply(lambda x4: "On" if x4==1 else "Off")
        df_final['RELAY5']=df_final['RELAY5'].apply(lambda x5: "On" if x5==1 else "Off")
        df_final['RELAY6']=df_final['RELAY6'].apply(lambda x6: "On" if x6==1 else "Off")
        df_final=df_final[['Time','pH','pH within specs','ORP (mV)','ORP within specs','CONDUCTIVITY '+ u'\u00B5' + "s/cm",'CONDUCTIVITY within specs',"Temp " + u'\u00B0'+"C",'CORROSION (MPY)','CORROSION within specs','FLUOROMETER (PPM)','FLUOROMETER within specs','Flowswitch','RELAY1','RELAY2','RELAY3','RELAY4','RELAY5','RELAY6']]

	
        df_final["pH within specs"] =df_final["pH within specs"].replace(np.nan, 'N')
        df_final["ORP within specs"] =df_final["ORP within specs"].replace(np.nan, 'N')
        df_final["CONDUCTIVITY within specs"] =df_final["CONDUCTIVITY within specs"].replace(np.nan, 'N')
        df_final["CORROSION within specs"]  =df_final["CORROSION within specs"].replace(np.nan, 'N')
        df_final["FLUOROMETER within specs"] =df_final["FLUOROMETER within specs"].replace(np.nan, 'N')
        df_final["Temp " + u'\u00B0'+"C"] =df_final["Temp " + u'\u00B0'+"C"].replace(np.nan, 'No Data')
        df_final['Time'] = pd.to_datetime(df_final['Time'])
        df_final['Time']=[datetime.fromtimestamp(x.value/1000) for x in df_final['Time']]
        

        df_final.replace(np.nan,"No Data",inplace=True)
        #print(df_final)
        FinalReport = "/space/es-master/src/Shravya/"+unitName + "_Report.xlsx"
        #df_final.to_excel(FinalReport,index=False)
        #return df_final

        df_final.to_excel(FinalReport,index=False,encoding='utf-8')
        wbook = Workbook()
        ws=wbook.active
        ws.title = "Sheet1"
        wbook.save(filename= FinalReport)
        wbook.close()

        writer = pd.ExcelWriter(FinalReport, engine='openpyxl')
        writer.book=load_workbook(FinalReport)
        df_final.to_excel(writer,index=False,encoding='utf-8')
        worksheet = writer.sheets["Sheet1"] # Access the Worksheet
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    try:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                    except:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(cell.value.encode('utf-8'))))
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value + 3
            
            
        worksheet.freeze_panes = 'B1'

        bd = Side(style='thin', color="000000")
        for cell in worksheet._cells.values():
            cell.font = Font(name='Verdana',size=10)
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1,max_col=df_final.shape[0]):
            for cell in row:
                cell.font = Font(name='Verdana',size=10,bold=True)
                cell.fill = PatternFill(fill_type='solid',start_color="00FF9900",end_color="00FF9900")

        writer.save()
        
        # wb = load_workbook(FinalReport)
        # if 'Sheet1' in wb.sheetnames:
            # wb.remove(wb['Sheet1'])
        # wb.save(FinalReport)

        # wb = load_workbook(FinalReport)
        # emails=["nitesh.tiwari@thermaxglobal.com","Vinod.desai@thermaxglobal.com"]
        # ccemail=["shravya.k@exactspace.co"]
 
        # body = {
                # "to": emails,
                # "cc": ccemail,
                # "subject": unitName+"-Daily Report",
                # "html":"<h3>Daily Report attached.</h3>",
                # "f1": FinalReport,
                # "f2":"", # Not mandatory
                # "f3":"" # Not mandatory
                # }
        # print("mailing")
        # email.sendSESMailWithAttach(body)
       


df=ExcelReader()